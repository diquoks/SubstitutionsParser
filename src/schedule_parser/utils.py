from __future__ import annotations

import datetime
import itertools
import typing

import openpyxl.cell.cell
import openpyxl.worksheet.worksheet
import pyquoks.models

import schedule_parser.constants
import schedule_parser.models


def _get_data_from_models_iterable(
        models_iterable: typing.Iterable[pyquoks.models._HasInitialData],
) -> list[typing.Any]:
    def _get_model_data(model: pyquoks.models._HasInitialData) -> typing.Any:
        return model._data

    return list(
        map(
            _get_model_data,
            models_iterable,
        )
    )


def parse_schedule(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[schedule_parser.models.GroupScheduleContainer]:
    def _get_periods(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[schedule_parser.models.PeriodModel]:
        def _get_number(row: int) -> int:
            return int(
                worksheet.cell(
                    row=row,
                    column=schedule_parser.constants.SCHEDULE_DATA_NUMBER_COLUMN,
                ).value
            )

        def _get_subgroup(column: int) -> int:
            return worksheet.cell(
                row=schedule_parser.constants.SCHEDULE_DATA_SUBGROUP_ROW,
                column=column,
            ).value

        split_columns = [
            list(
                map(
                    lambda batched: list(batched),
                    itertools.batched(
                        column,
                        schedule_parser.constants.SCHEDULE_PERIOD_HEIGHT,
                    ),
                )
            ) for column in columns
        ]

        for period_values in [list(current_columns) for current_columns in zip(*split_columns)]:
            if isinstance(period_values[1][0], openpyxl.cell.cell.MergedCell):
                yield schedule_parser.models.PeriodModel(
                    data={
                        "lecturer": period_values[0][1].value,
                        "number": _get_number(period_values[0][0].row),
                        "room": period_values[3][0].value,
                        "subgroup": schedule_parser.constants.SCHEDULE_DATA_NO_SUBGROUP,
                        "subject": period_values[0][0].value,
                    },
                )
            else:
                for first_column, second_column in [(0, 1), (2, 3)]:
                    if period_values[first_column][0].value:
                        yield schedule_parser.models.PeriodModel(
                            data={
                                "lecturer": period_values[first_column][1].value,
                                "number": _get_number(period_values[first_column][0].row),
                                "room": period_values[second_column][0].value,
                                "subgroup": _get_subgroup(period_values[first_column][0].column),
                                "subject": period_values[first_column][0].value,
                            },
                        )

    def _get_day_schedule(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[schedule_parser.models.DayScheduleContainer]:
        def _get_weekday(row: int) -> schedule_parser.models.Weekday:
            return schedule_parser.models.Weekday.from_string(
                string=worksheet.cell(
                    row=row,
                    column=schedule_parser.constants.SCHEDULE_DATA_WEEKDAY_COLUMN,
                ).value,
            )

        split_columns = []

        for column in columns:
            current_columns = []
            current_column = []

            for cell in column:
                current_column.append(cell)
                if cell.border.bottom.style == schedule_parser.constants.SCHEDULE_DATA_WEEKDAY_BORDER:
                    current_columns.append(current_column.copy())
                    current_column.clear()

            split_columns.append(current_columns)

        for day_schedule in [list(current_columns) for current_columns in zip(*split_columns)]:
            yield schedule_parser.models.DayScheduleContainer(
                data={
                    "weekday": _get_weekday(day_schedule[0][0].row).value,
                    "schedule": _get_data_from_models_iterable(
                        models_iterable=_get_periods(day_schedule),
                    ),
                },
            )

    def _get_week_schedule(
            columns: list[list[openpyxl.cell.cell.Cell]],
    ) -> typing.Generator[schedule_parser.models.WeekScheduleContainer]:
        def _check_cell_color(cell_index: int) -> bool:
            return columns[0][cell_index].fill.fgColor.rgb == schedule_parser.constants.SCHEDULE_DATA_ODD_COLOR

        week_variants = [
            (
                [
                    list(
                        filter(
                            lambda cell: _check_cell_color(column.index(cell)) == parity,
                            column,
                        )
                    ) for column in columns
                ],
                not parity,
            ) for parity in (True, False)
        ]

        for week_variant in week_variants:
            yield schedule_parser.models.WeekScheduleContainer(
                data={
                    "parity": week_variant[1],
                    "schedule": _get_data_from_models_iterable(
                        models_iterable=_get_day_schedule(week_variant[0]),
                    ),
                },
            )

    def _get_group(column: int) -> str:
        return worksheet.cell(
            row=schedule_parser.constants.SCHEDULE_DATA_GROUP_ROW,
            column=column,
        ).value

    for group_columns in itertools.batched(
            list(worksheet.columns)[
                schedule_parser.constants.SCHEDULE_SIDEBAR_WIDTH:schedule_parser.constants.SCHEDULE_JUNK_OFFSET
            ],
            schedule_parser.constants.SCHEDULE_GROUP_WIDTH,
    ):
        group_columns = list(
            map(
                lambda column: column[
                    schedule_parser.constants.SCHEDULE_HEADER_INDEX:schedule_parser.constants.SCHEDULE_FOOTER_OFFSET
                ],
                list(group_columns[-schedule_parser.constants.SCHEDULE_JUNK_OFFSET:]),
            )
        )

        yield schedule_parser.models.GroupScheduleContainer(
            data={
                "group": _get_group(group_columns[0][0].column),
                "schedule": _get_data_from_models_iterable(
                    models_iterable=_get_week_schedule(group_columns),
                ),
            },
        )


def parse_substitutions(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
) -> typing.Generator[schedule_parser.models.SubstitutionModel]:
    def _get_period(*args, number: int) -> schedule_parser.models.PeriodModel:
        return schedule_parser.models.PeriodModel(
            data={
                "lecturer": args[2],
                "number": number,
                "room": args[3],
                "subgroup": int(args[0]) if args[0] else 0,
                "subject": args[1],
            },
        )

    for row in list(worksheet.rows)[schedule_parser.constants.SUBSTITUTIONS_HEADER_INDEX:]:
        row_values = list(
            map(
                lambda cell: cell.value,
                list(row[:schedule_parser.constants.SUBSTITUTIONS_WIDTH]),
            )
        )

        if set(row_values) == {None}:
            continue

        group, period_number, *row_values = row_values
        period_number = int(period_number)

        yield schedule_parser.models.SubstitutionModel(
            data={
                "group": group,
                "period": _get_period(
                    *row_values[:4],
                    number=period_number,
                )._data,
                "substitution": _get_period(
                    *row_values[4:],
                    number=period_number,
                )._data,
            },
        )


def get_week_number(date: datetime.datetime) -> int:
    current_week_number = date.isocalendar().week

    first_school_week_number = datetime.datetime(
        year=date.year - 1 if current_week_number < datetime.datetime(
            year=date.year - 1,
            month=9,
            day=1,
        ).isocalendar().week else date.year,
        month=9,
        day=1,
    ).isocalendar().week

    last_year_week_number = datetime.datetime(
        year=date.year - 1 if current_week_number < datetime.datetime(
            year=date.year - 1,
            month=12,
            day=28,
        ).isocalendar().week else date.year,
        month=12,
        day=28,
    ).isocalendar().week

    if first_school_week_number <= current_week_number:
        return current_week_number - first_school_week_number + 1
    else:
        return last_year_week_number - first_school_week_number + current_week_number + 1


def get_schedule_with_substitutions(
        schedule: schedule_parser.models.GroupSchedulesListing,
        substitutions: schedule_parser.models.SubstitutionsListing,
        group: str,
        date: datetime.datetime,
) -> list[schedule_parser.models.PeriodModel]:
    schedule = (
        schedule.get_group_schedule(group)
        .get_week_schedule_by_parity(get_week_number(date) % 2 == 0)
        .get_day_schedule_by_weekday(schedule_parser.models.Weekday(date.weekday() + 1))
        .schedule
    )

    substitutions = substitutions.get_substitutions_by_group(group)

    new_schedule = schedule.copy()

    for substitution in substitutions:
        for index, period_model in enumerate(new_schedule):
            if period_model.is_same(substitution.period):
                new_schedule[index] = substitution.substitution
                break
        else:
            if not substitution.substitution.is_empty:
                new_schedule.append(substitution.substitution)

    new_schedule.sort(
        key=lambda period: (period.number, period.subgroup),
    )

    return list(
        filter(
            lambda period: not period.is_empty,
            new_schedule,
        )
    )
