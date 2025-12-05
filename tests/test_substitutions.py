from __future__ import annotations

import openpyxl
import pyquoks

import schedule_parser


class TestSubstitutions(pyquoks.test.TestCase):
    _MODULE_NAME = __name__

    def test_parse_substitutions(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions.xlsx"),
        )

        for substitution in list(schedule_parser.utils.parse_substitutions(workbook.worksheets[0])):
            self.assert_type(
                func_name=self.test_parse_substitutions.__name__,
                test_data=substitution,
                test_type=schedule_parser.models.SubstitutionModel,
            )

    def test_parse_substitutions_incorrect(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions_incorrect.xlsx"),
        )

        for substitution in list(schedule_parser.utils.parse_substitutions(workbook.worksheets[0])):
            self.assert_type(
                func_name=self.test_parse_substitutions_incorrect.__name__,
                test_data=substitution,
                test_type=schedule_parser.models.SubstitutionModel,
            )

    def test_parse_substitutions_old(self):
        workbook = openpyxl.load_workbook(
            filename=pyquoks.utils.get_path("resources/tables/substitutions_old.xlsx"),
        )

        for substitution in list(schedule_parser.utils.parse_substitutions(workbook.worksheets[0])):
            self.assert_type(
                func_name=self.test_parse_substitutions_old.__name__,
                test_data=substitution,
                test_type=schedule_parser.models.SubstitutionModel,
            )
